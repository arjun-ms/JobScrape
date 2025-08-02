# a Python script that scrapes job postings from a target company's public
# careers page and organizes the extracted information into a structured Excel file.

import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from datetime import datetime
import os

BASE_URL = "https://techversantinfotech.com/talent/"

def extract_job_data(job_section):
    job_data = {}

    try:
        # Primary fields - use empty string for missing data (not 'N/A')
        job_title_elem = job_section.find('h3', class_='crr_app_hh')
        job_data['JobTitle'] = job_title_elem.get_text(strip=True) if job_title_elem else ''
        
        job_category_elem = job_section.find('span', class_='crr_app_tp bluecrr')
        job_data['JobCategory'] = job_category_elem.get_text(strip=True) if job_category_elem else ''
        
        location_elem = job_section.find('span', class_='crr_app_plc')
        job_data['Location'] = location_elem.get_text(strip=True) if location_elem else ''

        # Apply button and job ID
        apply_button = job_section.find('a', class_='crr_app_nw')
        job_data['JobURL'] = apply_button.get('href') if apply_button else ''
        job_data['JobID'] = apply_button.get('datatitle') if apply_button else ''

        # Posting date
        posting_date_elem = job_section.find('p', string=lambda text: text and 'Posted on' in text)
        job_data['PostingDate'] = posting_date_elem.get_text(strip=True) if posting_date_elem else ''

        # Job Description Summary
        desc_section = job_section.find('strong', string='What\'s important to us:')
        if desc_section:
            desc_para = desc_section.parent.find_next_sibling('p')
            job_data['JobDescriptionSummary'] = desc_para.get_text(strip=True) if desc_para else ''
        else:
            job_data['JobDescriptionSummary'] = ''

        # Experience requirement extraction
        full_text = job_section.get_text()
        experience_patterns = [
            r'(\d+\+?\s*years?\s*of\s*experience)',
            r'(Minimum\s*of\s*\d+\s*years)',
            r'(\d+\+\s*years?\s*in)',
            r'(minimum\s*\d+\s*years)'
        ]
        
        experience_found = None
        for pattern in experience_patterns:
            experience_match = re.search(pattern, full_text, re.IGNORECASE)
            if experience_match:
                experience_found = experience_match.group(0)
                break
        
        # Leave blank if no experience found (not 'N/A')
        job_data['ExperienceRequired'] = experience_found if experience_found else ''

        # Skills extraction
        skills_sections = job_section.find_all('strong', string=lambda text: text and any(
            keyword in text for keyword in ['Preferred Skills', 'Required Skills', 'Must-Have Skills', 'Technical Stack']
        ))
        
        all_skills = []
        for skills_section in skills_sections:
            skills_list = skills_section.parent.find_next_sibling('ul')
            if skills_list:
                skills = [li.get_text(strip=True) for li in skills_list.find_all('li')]
                all_skills.extend(skills)
        
        job_data['SkillsRequired'] = '; '.join(all_skills) if all_skills else ''

        # Contact email
        email_link = job_section.find('a', href=lambda href: href and 'mailto:' in href)
        job_data['ContactEmail'] = email_link.get('href').replace('mailto:', '') if email_link else ''

        # Company benefits
        benefits_section = job_section.find('strong', string='What Company Offers:')
        if benefits_section:
            benefits_text = benefits_section.parent.find_next_sibling('p')
            job_data['CompanyBenefits'] = benefits_text.get_text(strip=True) if benefits_text else ''
        else:
            job_data['CompanyBenefits'] = ''

        # Salary - leave blank as it's typically not specified
        job_data['Salary'] = ''

        # Add scraping metadata
        job_data['ScrapedDate'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        job_data['SourceURL'] = BASE_URL

    except AttributeError as e:
        print(f"⚠️  HTML structure issue for a job posting: {str(e)}")
        # Return None to skip this job rather than error data
        return None
    except Exception as e:
        print(f"❌ Error extracting job data: {str(e)}")
        return None

    return job_data

def check_pagination(soup):
    # Look for common pagination patterns
    pagination_indicators = [
        soup.find('div', class_='pagination'),
        soup.find('nav', class_='pagination'), 
        soup.find('div', class_='pager'),
        soup.find('ul', class_='pagination'),
        soup.find('a', string=lambda text: text and any(word in text.lower() for word in ['next', 'more', '»', '>', 'page'])),
        soup.find('button', string=lambda text: text and 'load more' in text.lower()),
    ]
    
    has_pagination = any(indicator for indicator in pagination_indicators)
    return has_pagination

def save_jobs_to_excel(jobs_data, filename=None):
    if not jobs_data:
        raise ValueError("No job data provided to save")
    
    # Generate filename if not provided
    if filename is None:
        # timestamp = datetime.now().strftime('%d-%m-%Y-%I-%M-%S-%p')
        filename = f"Techversant_Jobs.xlsx"
    
    # Ensure .xlsx extension
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    try:
        # Create DataFrame by converting list of dicts to DataFrame
        df = pd.DataFrame(jobs_data)
        
        # Reorder columns for better readability
        column_order = [
            'JobTitle', 'JobCategory', 'Location', 'ExperienceRequired', 
            'PostingDate', 'JobDescriptionSummary', 'SkillsRequired', 
            'ContactEmail', 'CompanyBenefits', 'Salary', 'JobURL', 
            'JobID', 'ScrapedDate', 'SourceURL'
        ]
        
        # Keep only columns that exist in the data
        existing_columns = [col for col in column_order if col in df.columns]
        df = df[existing_columns]
        
        # Replace empty strings with None for proper blank cells in Excel
        df = df.replace('', None)
        
        # Create Excel writer object
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Write main data to 'Jobs' sheet
            # index=False removes pandas index column (meets acceptance criteria)
            df.to_excel(writer, sheet_name='Jobs', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Jobs']
            
            # Format the Excel file
            format_excel_worksheet(worksheet, df)
            
            # Create summary sheet
            create_summary_sheet(workbook, df)
        
        print(f"✅ Successfully saved {len(jobs_data)} jobs to '{filename}'")
        return filename
        
    except Exception as e:
        print(f"❌ Error saving to Excel: {str(e)}")
        raise

def format_excel_worksheet(worksheet, df):  
    # Define styles
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format headers
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Auto-adjust column widths and apply borders
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            cell.border = border
            if cell.row > 1:  # Data rows
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set column width with limits
        adjusted_width = min(max(max_length + 2, 15), 80)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Set row height for better readability
    for row in range(2, len(df) + 2):
        worksheet.row_dimensions[row].height = 60

def create_summary_sheet(workbook, df):
    
    summary_sheet = workbook.create_sheet(title='Summary')
    
    # Summary statistics
    summary_data = [
        ['Metric', 'Value'],
        ['Total Jobs', len(df)],
        ['Scraping Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Source', 'Techversant Infotech'],
        ['Source URL', BASE_URL],
        [''],
        ['Jobs by Category', ''],
    ]
    
    # Add category breakdown
    if 'JobCategory' in df.columns:
        category_counts = df['JobCategory'].value_counts()
        for category, count in category_counts.items():
            summary_data.append([f'  {category}', count])
    
    summary_data.extend([
        [''],
        ['Jobs by Location', ''],
    ])
    
    # Add location breakdown
    if 'Location' in df.columns:
        location_counts = df['Location'].value_counts()
        for location, count in location_counts.items():
            summary_data.append([f'  {location}', count])
    
    # Write summary data
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Format headers
            if row_idx == 1 or (len(row_data) == 2 and row_data[1] == ''):
                cell.font = Font(bold=True)
                if row_idx == 1:
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
    
    # Auto-adjust column widths
    for column in summary_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = max(max_length + 2, 15)
        summary_sheet.column_dimensions[column_letter].width = adjusted_width

def validate_scraped_data(jobs_data):
    if not jobs_data:
        print("⚠️  No jobs found - this might indicate a problem")
        return False
    
    # Check if we have essential fields
    essential_fields = ['JobTitle', 'JobCategory', 'Location']
    complete_jobs = 0
    
    for job in jobs_data:
        if all(job.get(field) for field in essential_fields):
            complete_jobs += 1
    
    completion_rate = (complete_jobs / len(jobs_data)) * 100
    print(f"📊 Data completeness: {complete_jobs}/{len(jobs_data)} jobs ({completion_rate:.1f}%) have all essential fields")
    
    if completion_rate < 80:
        print("⚠️  Low completion rate - website structure may have changed")
        return False
    
    return True

def scrape_jobs():
    
    try:
        print("🔍 Starting job scraping from Techversant...")
        
        # Make request to the careers page
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        response = requests.get(BASE_URL, headers=headers, timeout=30)
        response.raise_for_status()
        
        # Parse HTML
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Check for pagination (acceptance criteria requirement)
        has_pagination = check_pagination(soup)
        if has_pagination:
            print("📄 Pagination detected - currently scraping first page only")
            print("💡 Future versions could implement multi-page scraping")
        else:
            print("📄 No pagination detected - scraping all available jobs")
        
        # Find all job sections
        job_sections = soup.find_all('section', class_='crr_app_stt')
        
        if not job_sections:
            print("❌ No job sections found. The website structure might have changed.")
            return None
        
        print(f"📋 Found {len(job_sections)} job postings")
        
        # Extract data from each job (store as dictionaries, append to list)
        jobs_data = []  # List to store job dictionaries
        successful_extractions = 0

        for i, job_section in enumerate(job_sections, 1):
            print(f"📝 Processing job {i}/{len(job_sections)}...")
            # Extract job data and return as dictionary
            job_data = extract_job_data(job_section)
            
            # Only append if extraction was successful (not None)
            if job_data:
                jobs_data.append(job_data)  # Append dictionary to list
                successful_extractions += 1
            else:
                print(f"⚠️  Skipped job {i} due to extraction error")
        
        print(f"✅ Successfully extracted {successful_extractions}/{len(job_sections)} jobs")
        
        # Validate data quality
        if not validate_scraped_data(jobs_data):
            print("⚠️  Data validation concerns detected")
        
        # Save to Excel (convert list of dictionaries to DataFrame)
        if jobs_data:
            filename = save_jobs_to_excel(jobs_data)
            print(f"\n✅ Scraping completed successfully!")
            print(f"📁 File saved as: {filename}")
            print(f"📊 Total jobs scraped: {len(jobs_data)}")
            return filename
        else:
            print("❌ No job data extracted")
            return None
            
    except requests.RequestException as e:
        print(f"❌ Error fetching webpage: {str(e)}")
        print("💡 Check your internet connection and the website URL")
        return None
    except Exception as e:
        print(f"❌ Unexpected error: {str(e)}")
        return None

if __name__ == "__main__":
    # Run the scraping
    result = scrape_jobs()
    
    if result:
        print(f"\n🎉 Job scraping completed! Check '{result}' for results.")
    else:
        print("\n💥 Job scraping failed. Please check the errors above.")